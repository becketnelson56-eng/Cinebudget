"""
Standalone test for excel_writer.py using mock data.
Verifies formulas, hyperlinks, and structure without needing the Claude API.
"""
import pathlib
from openpyxl import load_workbook
from excel_writer import write_breakdown, DEPARTMENTS

MOCK_PDF = str(pathlib.Path("mock_script.pdf").resolve())

MOCK_ITEMS = [
    {"department": "Grip & Electric", "scene_number": "1", "item_name": "Fog machine + fluid",
     "description": "Practical fog effect for night plaza scene.", "quantity": 2,
     "unit_type": "Rental-Day", "script_page": 2, "char_offset": 120,
     "script_quote": "A thick mist hangs low over the cobblestones of CITY PLAZA",
     "notes": "Check if VFX if supernatural", "confidence": "High"},
    {"department": "Grip & Electric", "scene_number": "1", "item_name": "Rain tower / rain bar kit",
     "description": "Practical rain rig for wet street effect.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 2, "char_offset": 200,
     "script_quote": "The streets glisten with the remnants of a downpour",
     "notes": "Coordinate with Locations re: drainage", "confidence": "High"},
    {"department": "Grip & Electric", "scene_number": "2", "item_name": "Night lighting package",
     "description": "Additional lighting cost for night exterior scenes.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 2, "char_offset": 50,
     "script_quote": "EXT. CITY PLAZA - NIGHT",
     "notes": "Applies to all NIGHT scenes", "confidence": "High"},
    {"department": "Grip & Electric", "scene_number": "3", "item_name": "LED volume / projection rig",
     "description": "LED wall for in-camera city skyline projection.", "quantity": 1,
     "unit_type": "Rental-Week", "script_page": 3, "char_offset": 900,
     "script_quote": "a massive LED WALL cycles through a projected city skyline",
     "notes": "High-cost item - flag for producer review", "confidence": "High"},
    {"department": "Cast & Extras", "scene_number": "1", "item_name": "Background extras",
     "description": "Large crowd for City Plaza scene.", "quantity": 200,
     "unit_type": "Rental-Day", "script_page": 2, "char_offset": 350,
     "script_quote": "Hundreds more fill the plaza behind them barely visible through the fog",
     "notes": "Quantity flagged Low confidence", "confidence": "Medium"},
    {"department": "Miscellaneous", "scene_number": "1", "item_name": "Animal + licensed handler (dog)",
     "description": "German Shepherd and licensed wrangler.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 2, "char_offset": 500,
     "script_quote": "A GERMAN SHEPHERD trots to Marlowe side, sits, and looks up",
     "notes": "Licensed wrangler legally required", "confidence": "High"},
    {"department": "Props", "scene_number": "3", "item_name": "Rotary phone",
     "description": "Period-appropriate rotary phone for hotel bar counter.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 3, "char_offset": 800,
     "script_quote": "a rotary phone on the counter",
     "notes": "", "confidence": "High"},
    {"department": "Props", "scene_number": "5", "item_name": "Prop pistol (licensed armourer)",
     "description": "Prop firearm for assassin car chase sequence.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 5, "char_offset": 1850,
     "script_quote": "the assassin leans from the sedan window and FIRES a pistol",
     "notes": "Licensed armourer legally required on set", "confidence": "High"},
    {"department": "VFX", "scene_number": "3", "item_name": "VFX supervision (LED wall)",
     "description": "VFX supervisor for LED wall plate integration.", "quantity": 1,
     "unit_type": "Rental-Week", "script_page": 3, "char_offset": 900,
     "script_quote": "a massive LED WALL cycles through a projected city skyline",
     "notes": "", "confidence": "High"},
    {"department": "Camera", "scene_number": "4", "item_name": "Drone unit + FAA-licensed operator",
     "description": "Aerial drone shot over rooftop.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 4, "char_offset": 1200,
     "script_quote": "A DRONE SHOT sweeps over the rooftop revealing the full scale",
     "notes": "Separate line from main camera package", "confidence": "High"},
    {"department": "Stunts & Safety", "scene_number": "4", "item_name": "Stunt coordinator + performer (fall)",
     "description": "Three-story stunt fall off hotel rooftop.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 4, "char_offset": 1350,
     "script_quote": "himself off the edge -- a three-story free fall, arms spread",
     "notes": "", "confidence": "High"},
    {"department": "Stunts & Safety", "scene_number": "5", "item_name": "Stunt driver + coordinator",
     "description": "Stunt driver for car chase sequence.", "quantity": 2,
     "unit_type": "Rental-Day", "script_page": 5, "char_offset": 1600,
     "script_quote": "A BLACK SEDAN tears around the corner fishtailing on the wet cobblestones",
     "notes": "", "confidence": "High"},
    {"department": "Stunts & Safety", "scene_number": "5", "item_name": "On-set safety officer (armourer)",
     "description": "Licensed armourer for prop firearm.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 5, "char_offset": 1850,
     "script_quote": "the assassin leans from the sedan window and FIRES a pistol",
     "notes": "", "confidence": "High"},
    {"department": "Stunts & Safety", "scene_number": "6", "item_name": "Practical explosion + safety coordinator",
     "description": "Controlled explosion for wall demolition sequence.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 6, "char_offset": 2100,
     "script_quote": "an EXPLOSION tears through the wall behind the assassin",
     "notes": "May also need VFX line item", "confidence": "High"},
    {"department": "Transportation", "scene_number": "5", "item_name": "Picture vehicles (sedan + cruiser)",
     "description": "Two period picture vehicles for car chase.", "quantity": 2,
     "unit_type": "Rental-Day", "script_page": 5, "char_offset": 1600,
     "script_quote": "A BLACK SEDAN tears around the corner fishtailing on the wet cobblestones",
     "notes": "", "confidence": "High"},
    {"department": "Art Dept", "scene_number": "6", "item_name": "Artificial snow / ice materials",
     "description": "Artificial snow effect for alley climax.", "quantity": 1,
     "unit_type": "Purchase", "script_page": 6, "char_offset": 2250,
     "script_quote": "Snow begins to fall -- thick, artificial-looking flakes that dust the alley",
     "notes": "Flag Locations for weather day", "confidence": "High"},
    {"department": "Art Dept", "scene_number": "1", "item_name": "Period set dressing (1947)",
     "description": "1940s era set dressing for all locations.", "quantity": 1,
     "unit_type": "Rental-Week", "script_page": 2, "char_offset": 180,
     "script_quote": "It is 1947. The streets glisten with the remnants of a downpour.",
     "notes": "", "confidence": "Medium"},
    {"department": "Wardrobe", "scene_number": "1", "item_name": "Period costumes (1947)",
     "description": "1940s wardrobe for all principal cast and extras.", "quantity": 1,
     "unit_type": "Rental-Week", "script_page": 2, "char_offset": 180,
     "script_quote": "It is 1947. The streets glisten with the remnants of a downpour.",
     "notes": "Applies to all scenes", "confidence": "Medium"},
    {"department": "Locations", "scene_number": "1", "item_name": "Location permit + fees (City Plaza)",
     "description": "Permit for exterior public plaza filming.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 2, "char_offset": 50,
     "script_quote": "EXT. CITY PLAZA - NIGHT",
     "notes": "Every exterior public location needs permit", "confidence": "High"},
    {"department": "Catering", "scene_number": "3", "item_name": "On-set catering",
     "description": "Catering for cast and crew.", "quantity": 1,
     "unit_type": "Rental-Day", "script_page": 3, "char_offset": 1050,
     "script_quote": "catering trays are stacked near the service door",
     "notes": "", "confidence": "Medium"},
]


def test():
    output = "test_output.xlsx"
    print(f"Writing {len(MOCK_ITEMS)} mock items to {output}...")
    write_breakdown(MOCK_ITEMS, output, MOCK_PDF)
    print("Written. Verifying...")

    wb = load_workbook(output)

    # 1. Check all 14 tabs exist (Summary + 13 depts)
    expected_tabs = ["Summary"] + DEPARTMENTS
    for tab in expected_tabs:
        assert tab in wb.sheetnames, f"MISSING TAB: {tab}"
    print(f"  [OK] All {len(expected_tabs)} tabs present: {wb.sheetnames}")

    # 2. Verify Summary is first
    assert wb.sheetnames[0] == "Summary", "Summary tab must be first"
    print("  [OK] Summary tab is first")

    # 3. Check column G formulas in each dept tab that has data
    tabs_with_data = []
    for dept in DEPARTMENTS:
        ws = wb[dept]
        rows_with_data = [r for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value]
        for r in rows_with_data:
            expected_formula = f"=D{r}*F{r}"
            actual = ws.cell(r, 7).value
            assert actual == expected_formula, (
                f"[{dept}] Row {r} col G: expected '{expected_formula}', got '{actual}'"
            )
        if rows_with_data:
            tabs_with_data.append(dept)
    print(f"  [OK] Column G formula =Dn*Fn verified in {len(tabs_with_data)} dept tabs")

    # 4. Verify F, H, I are blank in data rows
    for dept in tabs_with_data:
        ws = wb[dept]
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 2).value:  # if row has an item name
                for col, name in [(6, "F (Unit Price)"), (8, "H (Vendor)"), (9, "I (Vendor URL)")]:
                    val = ws.cell(r, col).value
                    assert val is None, f"[{dept}] Row {r} col {name} should be blank, got: {val}"
    print("  [OK] Columns F, H, I are blank in all data rows")

    # 5. Check Summary formulas
    ws_sum = wb["Summary"]
    found_counta = False
    found_sum = False
    found_grand = False
    for row in ws_sum.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if "COUNTA" in cell.value:
                    found_counta = True
                if cell.value.startswith("=SUM(") and "!" in cell.value:
                    found_sum = True
                if cell.value.startswith("=SUM(C") and "!" not in cell.value:
                    found_grand = True
    assert found_counta, "Summary tab missing COUNTA formula"
    assert found_sum, "Summary tab missing SUM dept subtotal formula"
    assert found_grand, "Summary tab missing grand total SUM formula"
    print("  [OK] Summary tab has COUNTA, SUM subtotals, and grand total formulas")

    # 6. Check column J hyperlinks
    hyperlink_count = 0
    for dept in tabs_with_data:
        ws = wb[dept]
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 2).value:
                cell_j = ws.cell(r, 10)
                if cell_j.value is not None and cell_j.hyperlink:
                    hyperlink_count += 1
    print(f"  [OK] Column J: {hyperlink_count} hyperlinks found")

    # 7. Confirm no hardcoded numbers in formula cells
    for dept in tabs_with_data:
        ws = wb[dept]
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 2).value:
                g_val = ws.cell(r, 7).value
                assert isinstance(g_val, str) and g_val.startswith("="), (
                    f"[{dept}] Row {r} col G is not a formula string: {g_val!r}"
                )
    print("  [OK] No hardcoded values in column G — all formula strings")

    print(f"\nAll checks passed. Spreadsheet: {output}")
    print("\nColumn layout (Tool 2 integration contract):")
    print("  F = Unit Price  (Tool 2 writes here)")
    print("  H = Vendor      (Tool 2 writes here)")
    print("  I = Vendor URL  (Tool 2 writes here)")


if __name__ == "__main__":
    test()

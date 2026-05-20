import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# ── Style constants ───────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BASE_FONT   = Font(size=10)

# Column O (Populated By) fill per type
FILL_SCRIPT   = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
FILL_SCHEDULE = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
FILL_DEAL     = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
FILL_FORMULA  = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
FILL_TEMPLATE = PatternFill(start_color="F1F3F5", end_color="F1F3F5", fill_type="solid")
FILL_WHITE    = PatternFill(fill_type=None)

COL_O_FILL = {
    "SCRIPT": FILL_SCRIPT,
    "SCHEDULE": FILL_SCHEDULE,
    "DEAL": FILL_DEAL,
    "FORMULA": FILL_FORMULA,
}

FORMULA_FONT    = Font(color="385723", size=10)
HYPERLINK_FONT  = Font(color="0563C1", underline="single", size=10)
PLACEHOLDER_FONT = Font(color="6C757D", italic=True, size=10)

CONF_FONT = {
    "High":   Font(color="375623", bold=True, size=10),
    "Medium": Font(color="974706", bold=True, size=10),
    "Low":    Font(color="9C0006", bold=True, size=10),
}

WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")
TOP_LEFT = Alignment(vertical="top")

# ── Account definitions ───────────────────────────────────────────────────────

ACCOUNTS: list[tuple[str, str]] = [
    ("1100", "Development"),
    ("1200", "Story & Rights"),
    ("1400", "Producers Unit"),
    ("1500", "Directors Unit"),
    ("1600", "Talent (Cast)"),
    ("2100", "Production Staff"),
    ("2200", "Art Direction"),
    ("2300", "Set Construction"),
    ("2400", "Set Decoration"),
    ("2500", "Props"),
    ("2600", "Camera"),
    ("2700", "Electric"),
    ("2800", "Grip"),
    ("2900", "Production Sound"),
    ("2950", "Stunts & Safety"),
    ("3000", "Mechanical FX"),
    ("3100", "Special VFX"),
    ("3200", "Set Operations"),
    ("3300", "Wardrobe"),
    ("3400", "Makeup & Hair"),
    ("3500", "Locations"),
    ("3600", "Transportation"),
    ("3800", "Studio Facilities"),
    ("3900", "Atmosphere"),
    ("5100", "Editing"),
    ("5200", "Post Film & Lab"),
    ("5300", "Post Sound"),
    ("5400", "Music"),
    ("5500", "Titles"),
    ("5600", "Opticals & Post VFX"),
    ("5700", "Post Video"),
    ("6100", "Insurance"),
    ("6200", "Legal"),
    ("6300", "Publicity"),
    ("6400", "Miscellaneous"),
]

# Standard Movie Magic template rows: (description, placeholder_type)
ACCOUNT_TEMPLATES: dict[str, list[tuple[str, str]]] = {
    "1100": [
        ("Story option / rights acquisition", "DEAL"),
        ("Screenplay development fee", "DEAL"),
        ("Development legal", "DEAL"),
        ("Development overhead", "DEAL"),
    ],
    "1200": [
        ("Story rights purchase", "DEAL"),
        ("Writer fee", "DEAL"),
        ("Continuity / breakdown", "DEAL"),
        ("Script duplication", "DEAL"),
    ],
    "1400": [
        ("Executive Producer fee", "DEAL"),
        ("Producer fee", "DEAL"),
        ("Line Producer fee", "DEAL"),
        ("Producer assistant(s)", "DEAL"),
    ],
    "1500": [
        ("Director fee", "DEAL"),
        ("Director assistant", "DEAL"),
        ("Storyboard artist", "DEAL"),
    ],
    "1600": [
        ("Casting Director", "DEAL"),
        ("Casting Associate", "DEAL"),
    ],
    "2100": [
        ("Production Manager / UPM", "SCHEDULE"),
        ("1st Assistant Director", "SCHEDULE"),
        ("2nd Assistant Director", "SCHEDULE"),
        ("Production Coordinator", "SCHEDULE"),
        ("Production Secretary", "SCHEDULE"),
        ("Location Manager", "SCHEDULE"),
        ("Production Accountant", "SCHEDULE"),
        ("Payroll services", "DEAL"),
    ],
    "2200": [
        ("Production Designer", "DEAL"),
        ("Art Director", "DEAL"),
        ("Set Designer / Drafting", "DEAL"),
        ("Art Department Assistant", "DEAL"),
        ("Art Department supplies", "DEAL"),
    ],
    "2300": [
        ("Construction Coordinator", "DEAL"),
        ("Construction labor", "DEAL"),
        ("Construction materials", "DEAL"),
        ("Stage / mill rental", "DEAL"),
    ],
    "2400": [
        ("Set Decorator", "DEAL"),
        ("Swing Gang", "DEAL"),
        ("Set dressings — purchased", "DEAL"),
        ("Set dressings — rented", "DEAL"),
        ("Graphics / signage", "DEAL"),
    ],
    "2500": [
        ("Property Master", "DEAL"),
        ("Assistant Property Master", "DEAL"),
        ("Prop rentals", "DEAL"),
        ("Prop purchases", "DEAL"),
    ],
    "2600": [
        ("Director of Photography", "DEAL"),
        ("Camera Operator", "DEAL"),
        ("1st AC / Focus Puller", "DEAL"),
        ("2nd AC / Loader", "DEAL"),
        ("Main Camera Package", "DEAL"),
    ],
    "2700": [
        ("Gaffer", "DEAL"),
        ("Best Boy Electric", "DEAL"),
        ("Electricians", "DEAL"),
        ("Lighting Package", "DEAL"),
        ("Generator rental", "DEAL"),
    ],
    "2800": [
        ("Key Grip", "DEAL"),
        ("Best Boy Grip", "DEAL"),
        ("Dolly Grip", "DEAL"),
        ("Grip Package", "DEAL"),
    ],
    "2900": [
        ("Production Sound Mixer", "DEAL"),
        ("Boom Operator", "DEAL"),
        ("Sound Package", "DEAL"),
        ("Walkie-talkies", "DEAL"),
    ],
    "2950": [
        ("Stunt Coordinator", "DEAL"),
        ("Stunt Performer(s)", "DEAL"),
        ("Safety Officer", "DEAL"),
        ("Stunt Equipment / Rigging", "DEAL"),
    ],
    "3000": [
        ("FX Coordinator", "DEAL"),
        ("FX Crew", "DEAL"),
        ("FX Materials / consumables", "DEAL"),
    ],
    "3100": [
        ("VFX Supervisor", "DEAL"),
        ("VFX Production / labor", "DEAL"),
        ("VFX Software / tools", "DEAL"),
    ],
    "3200": [
        ("Set Carpenter (standby)", "SCHEDULE"),
        ("Standby Painter", "SCHEDULE"),
        ("Craft Service", "SCHEDULE"),
        ("First Aid / On-set Nurse", "SCHEDULE"),
        ("Security", "SCHEDULE"),
    ],
    "3300": [
        ("Costume Designer", "DEAL"),
        ("Wardrobe Supervisor", "DEAL"),
        ("Costumer(s)", "DEAL"),
        ("Wardrobe Rental", "DEAL"),
        ("Wardrobe Purchases", "DEAL"),
        ("Dry cleaning / laundry", "DEAL"),
    ],
    "3400": [
        ("Head Makeup Artist", "DEAL"),
        ("Key Hair Stylist", "DEAL"),
        ("Additional Makeup / Hair", "DEAL"),
        ("Makeup Supplies", "DEAL"),
    ],
    "3500": [
        ("Location Manager", "DEAL"),
        ("Location Scout / Survey", "DEAL"),
        ("Police / fire standby", "DEAL"),
        ("Catering", "DEAL"),
        ("Parking / logistics", "DEAL"),
    ],
    "3600": [
        ("Transportation Coordinator", "SCHEDULE"),
        ("Captain Driver", "SCHEDULE"),
        ("Drivers", "SCHEDULE"),
        ("Production vehicles (trucks / vans)", "DEAL"),
        ("Honeywagon / basecamp", "DEAL"),
    ],
    "3800": [
        ("Stage rental", "DEAL"),
        ("Back lot rental", "DEAL"),
        ("Electric power (stage)", "DEAL"),
        ("Heating / AC", "DEAL"),
    ],
    "3900": [
        ("Extras Casting fee", "DEAL"),
        ("Stand-ins", "SCHEDULE"),
        ("Welfare Worker / Teacher (minors)", "DEAL"),
    ],
    "5100": [
        ("Post Production Supervisor", "DEAL"),
        ("Editor", "DEAL"),
        ("Assistant Editor(s)", "DEAL"),
        ("Editing Room / Avid rental", "DEAL"),
    ],
    "5200": [
        ("Color grading & Digital Intermediate", "DEAL"),
        ("Digital deliverables", "DEAL"),
        ("Archival / backup", "DEAL"),
    ],
    "5300": [
        ("Sound Designer", "DEAL"),
        ("Foley Artist + Studio", "DEAL"),
        ("ADR / Looping", "DEAL"),
        ("Final Mix", "DEAL"),
    ],
    "5400": [
        ("Composer fee", "DEAL"),
        ("Musicians / orchestration", "DEAL"),
        ("Music rights / licensing", "DEAL"),
        ("Score recording studio", "DEAL"),
    ],
    "5500": [
        ("Main titles design", "DEAL"),
        ("End credits", "DEAL"),
        ("Title card mastering", "DEAL"),
    ],
    "5600": [
        ("Opticals / visual effects (post)", "DEAL"),
        ("Digital intermediate (opticals)", "DEAL"),
    ],
    "5700": [
        ("Video deliverables / mastering", "DEAL"),
        ("Streaming / broadcast specs", "DEAL"),
    ],
    "6100": [
        ("Cast insurance", "DEAL"),
        ("Negative / production insurance", "DEAL"),
        ("E&O insurance", "DEAL"),
        ("General liability", "DEAL"),
    ],
    "6200": [
        ("Legal fees", "DEAL"),
        ("Contract preparation", "DEAL"),
        ("Clearances / chain of title", "DEAL"),
    ],
    "6300": [
        ("Unit Publicist", "DEAL"),
        ("Electronic Press Kit (EPK)", "DEAL"),
        ("Stills photographer", "DEAL"),
    ],
    "6400": [
        ("Accounting / payroll fees", "DEAL"),
        ("Bank charges", "DEAL"),
        ("Completion bond (~2.5% of budget)", "DEAL"),
    ],
}

# ── Account tab column spec ───────────────────────────────────────────────────

ACCT_HEADERS = [
    ("Account No",    10),
    ("Description",   40),
    ("Amount",        10),
    ("Unit Type",     14),
    ("Rate",          12),
    ("Subtotal",      12),
    ("Fringe %",      10),
    ("Total",         12),
    ("Vendor",        20),
    ("Vendor URL",    28),
    ("Script Page",   11),
    ("Script Quote",  35),
    ("Notes",         35),
    ("Confidence",    12),
    ("Populated By",  14),
]  # 15 columns A–O

# ── Scene Breakdown tab column spec ──────────────────────────────────────────

# (header_label, field_key_or_None, fill_hex_or_None)
SCENE_COLS: list[tuple[str, str | None, str | None, int]] = [
    ("Scene #",              "scene_number",       None,     8),
    ("Scene Heading",        "heading",            None,     32),
    ("Int/Ext",              "int_ext",            None,     8),
    ("Day/Night",            "day_night",          None,     10),
    ("Page Count",           "page_count_eighths", None,     10),
    ("Script Day",           "script_day",         None,     10),
    ("Cast",                 "cast",               "C00000", 22),   # Red
    ("Extras / Background",  "extras",             "9C6500", 16),   # Amber
    ("Props",                "props",              "375623", 22),   # Green
    ("Practical FX",         "practical_fx",       "2F5597", 16),   # Blue
    ("Wardrobe",             "wardrobe_notes",     "833C00", 18),   # Orange
    ("Animals",              "animals",            "3E1153", 13),   # Purple
    ("Vehicles / Picture Cars", "picture_vehicles","843C0C", 20),   # Brown
    ("Special Makeup",       "special_makeup",     "843C78", 16),   # Pink
    ("Special Camera",       "special_camera",     "006E6F", 16),   # Teal
    ("Stunts",               "stunts",             "404040", 20),   # Dark gray
    ("VFX",                  "vfx",                "006B8A", 18),   # Cyan
    ("Location Type",        "location_type",      None,     14),
    ("Shooting Day",         None,                 None,     12),   # [SCHEDULE]
    ("Est. Shoot Hours",     None,                 None,     14),   # [SCHEDULE]
    ("Notes / Flags",        "flags",              None,     30),
    ("Script Page",          "script_page",        None,     11),
]  # 22 columns A–V

# ── Public API ────────────────────────────────────────────────────────────────


def write_breakdown(data: dict, output_path: str, pdf_path: str) -> None:
    """
    Write production breakdown data to a V2 Movie Magic-structured .xlsx file.

    Args:
        data:        dict with 'scene_breakdown' list and 'accounts' dict,
                     as returned by api_client.get_breakdown()
        output_path: destination .xlsx path
        pdf_path:    absolute path of the source PDF (used for hyperlinks)
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws_top = _create_top_sheet(wb)
    _write_scene_breakdown_tab(wb, data.get("scene_breakdown", []), pdf_path)

    acct_data = data.get("accounts", {})
    for acct_no, acct_name in ACCOUNTS:
        ws = _create_account_tab(wb, acct_no, acct_name)
        _write_account_rows(ws, acct_no, acct_data.get(acct_no, []), pdf_path)

    _populate_top_sheet(ws_top)

    wb.save(output_path)


# ── Internal helpers ──────────────────────────────────────────────────────────


def _formula_tab_ref(tab_name: str) -> str:
    """Quote tab name for use inside an Excel formula if needed."""
    needs_quoting = (
        tab_name[0].isdigit()
        or any(c in tab_name for c in (" ", "&", "'", "!", "[", "]", "/", "\\"))
    )
    if needs_quoting:
        return f"'{tab_name.replace(chr(39), chr(39)*2)}'"
    return tab_name


def _set_hyperlink(cell, pdf_path: str, page_num) -> None:
    """Add a file:// hyperlink to a PDF page on the given cell."""
    try:
        page_num = int(page_num)
    except (TypeError, ValueError):
        return
    cell.value = page_num
    uri = pathlib.Path(pdf_path).resolve().as_uri()
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=uri)
    cell.font = HYPERLINK_FONT


def _list_to_cell(value) -> str:
    """Convert a list (or scalar) to a newline-joined string for a cell."""
    if isinstance(value, list):
        return "\n".join(str(v) for v in value if v)
    return str(value) if value else ""


# ── Top Sheet ─────────────────────────────────────────────────────────────────

# Fixed row layout for the Top Sheet (built once, referenced by _populate_top_sheet)
_TOP_SECTION_TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_TOP_TOTAL_FILL         = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_TOP_GRAND_FILL         = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_CURRENCY_FMT = '"$"#,##0.00'

# Row index → (label, account_ref_string_or_None, is_total, is_section, is_grand)
_TOP_ROWS: list[tuple[int, str, str | None, bool, bool, bool]] = [
    # row, label, acct_ref, is_total, is_section, is_grand
    (3,  "ABOVE THE LINE",                                         None,    False, True,  False),
    (4,  "1100-1200 – Development + Story & Rights",
         "SUM('1100 – Development'!H:H,'1200 – Story & Rights'!H:H)",      False, False, False),
    (5,  "1400 – Producers Unit",           "SUM('1400 – Producers Unit'!H:H)",           False, False, False),
    (6,  "1500 – Directors Unit",           "SUM('1500 – Directors Unit'!H:H)",           False, False, False),
    (7,  "1600 – Talent (Cast)",            "SUM('1600 – Talent (Cast)'!H:H)",            False, False, False),
    (8,  "TOTAL ABOVE THE LINE",            "SUM(C4:C7)",                                 True,  False, False),
    (10, "BELOW THE LINE",                                         None,    False, True,  False),
    (11, "2100 – Production Staff",         "SUM('2100 – Production Staff'!H:H)",         False, False, False),
    (12, "2200-2400 – Art Direction + Set Construction + Set Decoration",
         "SUM('2200 – Art Direction'!H:H,"
         "'2300 – Set Construction'!H:H,'2400 – Set Decoration'!H:H)",      False, False, False),
    (13, "2500 – Props",                    "SUM('2500 – Props'!H:H)",                    False, False, False),
    (14, "2600 – Camera",                   "SUM('2600 – Camera'!H:H)",                   False, False, False),
    (15, "2700 – Electric",                 "SUM('2700 – Electric'!H:H)",                 False, False, False),
    (16, "2800 – Grip",                     "SUM('2800 – Grip'!H:H)",                     False, False, False),
    (17, "2900 – Production Sound",         "SUM('2900 – Production Sound'!H:H)",         False, False, False),
    (18, "2950 – Stunts & Safety",          "SUM('2950 – Stunts & Safety'!H:H)",          False, False, False),
    (19, "3000-3100 – Mechanical FX + Special VFX",
         "SUM('3000 – Mechanical FX'!H:H,'3100 – Special VFX'!H:H)",       False, False, False),
    (20, "3200 – Set Operations",           "SUM('3200 – Set Operations'!H:H)",           False, False, False),
    (21, "3300 – Wardrobe",                 "SUM('3300 – Wardrobe'!H:H)",                 False, False, False),
    (22, "3400 – Makeup & Hair",            "SUM('3400 – Makeup & Hair'!H:H)",            False, False, False),
    (23, "3500 – Locations",                "SUM('3500 – Locations'!H:H)",                False, False, False),
    (24, "3600 – Transportation",           "SUM('3600 – Transportation'!H:H)",           False, False, False),
    (25, "3800-3900 – Studio Facilities + Atmosphere",
         "SUM('3800 – Studio Facilities'!H:H,'3900 – Atmosphere'!H:H)",    False, False, False),
    (26, "TOTAL BELOW THE LINE",            "SUM(C11:C25)",                               True,  False, False),
    (28, "POST PRODUCTION",                                        None,    False, True,  False),
    (29, "5100-5700 – Post Production (Editing, Lab, Sound, Music, Titles, Opticals)",
         "SUM('5100 – Editing'!H:H,'5200 – Post Film & Lab'!H:H,"
         "'5300 – Post Sound'!H:H,'5400 – Music'!H:H,'5500 – Titles'!H:H,"
         "'5600 – Opticals & Post VFX'!H:H,'5700 – Post Video'!H:H)",      False, False, False),
    (30, "TOTAL POST PRODUCTION",           "SUM(C29)",                                   True,  False, False),
    (32, "OTHER",                                                  None,    False, True,  False),
    (33, "6100 – Insurance",                "SUM('6100 – Insurance'!H:H)",                False, False, False),
    (34, "6200 – Legal",                    "SUM('6200 – Legal'!H:H)",                    False, False, False),
    (35, "6300 – Publicity",                "SUM('6300 – Publicity'!H:H)",                False, False, False),
    (36, "6400 – Contingency (10%)",        "(C8+C26+C30)*0.10",                          False, False, False),
    (37, "GRAND TOTAL",                     "C8+C26+C30+C33+C34+C35+C36",                False, False, True),
]


def _create_top_sheet(wb: Workbook):
    """Create the Top Sheet tab at position 0. Returns the worksheet."""
    ws = wb.create_sheet("TOP SHEET", 0)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16

    # Title
    title = ws.cell(row=1, column=1, value="PRODUCTION BUDGET — TOP SHEET")
    title.font = Font(bold=True, size=14, color="1F3864")
    ws.merge_cells("A1:C1")

    # Column headers
    for col, label in enumerate(["Category", "Account(s)", "Total"], start=1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    return ws


def _populate_top_sheet(ws) -> None:
    """Fill formula rows into the Top Sheet. Called after all account tabs exist."""
    for row, label, formula_body, is_total, is_section, is_grand in _TOP_ROWS:
        label_cell = ws.cell(row=row, column=1, value=label)
        total_cell = ws.cell(row=row, column=3)

        if is_grand:
            label_cell.font = Font(bold=True, size=11, color="FFFFFF")
            label_cell.fill = _TOP_GRAND_FILL
            total_cell.fill = _TOP_GRAND_FILL
            ws.cell(row=row, column=2).fill = _TOP_GRAND_FILL
            total_cell.font = Font(bold=True, size=11, color="FFFFFF")
            if formula_body:
                total_cell.value = f"={formula_body}"
        elif is_total:
            label_cell.font = Font(bold=True, size=10)
            label_cell.fill = _TOP_TOTAL_FILL
            total_cell.fill = _TOP_TOTAL_FILL
            ws.cell(row=row, column=2).fill = _TOP_TOTAL_FILL
            total_cell.font = Font(bold=True, size=10, color="385723")
            if formula_body:
                total_cell.value = f"={formula_body}"
        elif is_section:
            label_cell.font = Font(bold=True, size=10, color="FFFFFF")
            label_cell.fill = _TOP_SECTION_TITLE_FILL
            ws.cell(row=row, column=2).fill = _TOP_SECTION_TITLE_FILL
            ws.cell(row=row, column=3).fill = _TOP_SECTION_TITLE_FILL
        else:
            label_cell.font = BASE_FONT
            if formula_body:
                total_cell.value = f"={formula_body}"
                total_cell.font = FORMULA_FONT

        if not is_section:
            total_cell.number_format = _CURRENCY_FMT


# ── Scene Breakdown tab ───────────────────────────────────────────────────────


def _write_scene_breakdown_tab(wb: Workbook, scenes: list[dict], pdf_path: str) -> None:
    """Create and populate the Scene Breakdown tab."""
    ws = wb.create_sheet("SCENE BREAKDOWN", 1)
    ws.freeze_panes = "A2"

    # Header row
    for col, (label, _field, fill_hex, width) in enumerate(SCENE_COLS, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.alignment = Alignment(horizontal="center", wrap_text=False)
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = width
        if fill_hex:
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
        else:
            c.font = HEADER_FONT
            c.fill = HEADER_FILL

    # Data rows
    for i, scene in enumerate(scenes):
        row = i + 2
        ws.row_dimensions[row].height = 45

        for col, (label, field_key, _fill, _w) in enumerate(SCENE_COLS, start=1):
            cell = ws.cell(row=row, column=col)
            cell.font = BASE_FONT
            cell.alignment = WRAP

            # Schedule placeholders
            if label in ("Shooting Day", "Est. Shoot Hours"):
                cell.value = "[SCHEDULE]"
                cell.font = PLACEHOLDER_FONT
                continue

            # Script Page hyperlink (last column)
            if field_key == "script_page":
                page = scene.get("script_page")
                if page is not None:
                    _set_hyperlink(cell, pdf_path, page)
                continue

            if field_key is None:
                continue

            # Extras: combine count + description
            if field_key == "extras":
                count = scene.get("extras_count") or 0
                desc = scene.get("extras_description") or ""
                if count or desc:
                    cell.value = f"{count}" + (f" — {desc}" if desc else "")
                continue

            value = scene.get(field_key)
            if isinstance(value, list):
                cell.value = _list_to_cell(value) or None
            else:
                cell.value = value or None


# ── Account tabs ──────────────────────────────────────────────────────────────


def _create_account_tab(wb: Workbook, acct_no: str, acct_name: str):
    """Create an account worksheet with a formatted header row."""
    ws = wb.create_sheet(f"{acct_no} – {acct_name}")

    for col, (label, width) in enumerate(ACCT_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    return ws


def _write_account_rows(
    ws, acct_no: str, script_items: list[dict], pdf_path: str
) -> None:
    """
    Write SCRIPT rows first, then gray template placeholder rows.

    SCRIPT rows get formulas in columns F and H.
    Template rows are left blank in F and H (no value to calculate yet).

    Patch 1 (Hereditary): template rows are renumbered from max(last_auto_idx,
    highest_explicit_sub + 1) so they never collide with explicit account numbers
    that Claude may have set on SCRIPT rows (e.g. 2500.17).
    """
    row = 2
    sub_idx = 1

    # — SCRIPT rows ————————————————————————————————
    # Always use the sequential sub_idx for column A — never the item's original
    # account_no from Claude. Post-processing may remove rows, leaving gaps in
    # Claude's original numbering. sub_idx produces gap-free sequential output.
    for item in script_items:
        _write_script_row(ws, row, acct_no, sub_idx, item, pdf_path)
        row += 1
        sub_idx += 1

    # — Template placeholder rows follow immediately after the last SCRIPT row ——
    for desc, placeholder_type in ACCOUNT_TEMPLATES.get(acct_no, []):
        _write_template_row(ws, row, acct_no, sub_idx, desc, placeholder_type)
        row += 1
        sub_idx += 1


def _write_script_row(
    ws, row: int, acct_no: str, idx: int, item: dict, pdf_path: str
) -> None:
    """Write a single SCRIPT-extracted row onto an account tab."""
    # A — Account No (always sequential — never use item's original account_no)
    ws.cell(row=row, column=1, value=f"{acct_no}.{idx:02d}").font = BASE_FONT

    # B — Description
    b = ws.cell(row=row, column=2, value=item.get("description") or "")
    b.font = BASE_FONT
    b.alignment = WRAP

    # C — Amount
    amt = item.get("amount")
    if amt is not None:
        try:
            amt = int(amt)
        except (TypeError, ValueError):
            amt = None
    unit_type = item.get("unit_type") or ""
    notes = item.get("notes") or ""
    if amt is None and unit_type == "Allow":
        amt = 1
    if amt is None and acct_no == "3500" and unit_type == "Rental-Day":
        amt = 1
        schedule_note = "[SCHEDULE] — update with actual shoot days once schedule is locked."
        notes = f"{notes} {schedule_note}".strip() if notes else schedule_note
    ws.cell(row=row, column=3, value=amt).font = BASE_FONT

    # D — Unit Type
    ws.cell(row=row, column=4, value=unit_type).font = BASE_FONT

    # E — Rate (blank — Tool 2 / LP fills)
    ws.cell(row=row, column=5, value=None).number_format = _CURRENCY_FMT

    # F — Subtotal formula
    f_cell = ws.cell(row=row, column=6)
    f_cell.value = f'=IFERROR(C{row}*D{row}*E{row},"")'
    f_cell.font = FORMULA_FONT
    f_cell.number_format = _CURRENCY_FMT

    # G — Fringe % (blank)
    ws.cell(row=row, column=7, value=None)

    # H — Total formula
    h_cell = ws.cell(row=row, column=8)
    h_cell.value = f'=IFERROR(F{row}+(F{row}*G{row}),"")'
    h_cell.font = FORMULA_FONT
    h_cell.number_format = _CURRENCY_FMT

    # I — Vendor (blank)
    ws.cell(row=row, column=9, value=None)

    # J — Vendor URL (blank)
    ws.cell(row=row, column=10, value=None)

    # K — Script Page (hyperlink)
    k_cell = ws.cell(row=row, column=11)
    k_cell.font = BASE_FONT
    page = item.get("script_page")
    if page is not None:
        _set_hyperlink(k_cell, pdf_path, page)

    # L — Script Quote
    l_cell = ws.cell(row=row, column=12, value=item.get("script_quote") or "")
    l_cell.font = Font(italic=True, size=10)
    l_cell.alignment = WRAP

    # M — Notes
    m_cell = ws.cell(row=row, column=13, value=notes)
    m_cell.font = BASE_FONT
    m_cell.alignment = WRAP

    # N — Confidence (color-coded)
    conf = item.get("confidence") or ""
    n_cell = ws.cell(row=row, column=14, value=conf)
    n_cell.font = CONF_FONT.get(conf, BASE_FONT)

    # O — Populated By (green fill for SCRIPT)
    o_cell = ws.cell(row=row, column=15, value="SCRIPT")
    o_cell.font = Font(bold=True, size=10, color="375623")
    o_cell.fill = FILL_SCRIPT
    o_cell.alignment = CENTER


def _write_template_row(
    ws, row: int, acct_no: str, idx: int, description: str, placeholder_type: str
) -> None:
    """Write a gray placeholder row for a standard MM budget line item."""
    placeholder_label = f"[{placeholder_type}]"
    o_fill = COL_O_FILL.get(placeholder_type, FILL_TEMPLATE)

    for col in range(1, 16):
        ws.cell(row=row, column=col).fill = FILL_TEMPLATE

    ws.cell(row=row, column=1, value=f"{acct_no}.{idx:02d}").font = PLACEHOLDER_FONT
    ws.cell(row=row, column=1).fill = FILL_TEMPLATE

    b = ws.cell(row=row, column=2, value=description)
    b.font = PLACEHOLDER_FONT
    b.fill = FILL_TEMPLATE
    b.alignment = WRAP

    c = ws.cell(row=row, column=3, value=placeholder_label)
    c.font = PLACEHOLDER_FONT
    c.fill = FILL_TEMPLATE

    # F and H intentionally left blank for template rows (nothing to calculate yet)

    o = ws.cell(row=row, column=15, value=placeholder_type)
    o.font = Font(size=10, color="6C757D")
    o.fill = o_fill
    o.alignment = CENTER
